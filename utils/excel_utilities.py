import openpyxl


class Excel_Methods:

    def __init__(self, filename, sheet):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook[sheet]

    def write_data(self, row_num, column_num, data):
        self.sheet.cell(row=row_num, column=column_num).value = data
        self.workbook.save(self.filename)

    def write_numbers(self, row_num, column_num, data):
        if data == 0 or data == '0':
            self.sheet.cell(row=row_num, column=column_num).value = data
        else:
            self.sheet.cell(row=row_num, column=column_num).value = int(data)
        self.workbook.save(self.filename)

    def read_data(self, row_num, column_num):
        return self.sheet.cell(row=row_num, column=column_num).value

    get_column_count = lambda self: self.sheet.max_column

    get_row_count = lambda self: self.sheet.max_row
